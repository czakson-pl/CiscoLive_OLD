import os.path
import json
from termcolor import colored
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import quote_sheetname
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

#helper functions

def print_colored(msg1, col1, text1, no1, no2=None, no3=None):
    if no3==None:
        if no2==None:
            print(colored(str(no1), 'yellow'), colored(msg1, col1), colored(text1, 'white'))
        else:
            print(colored(str(no1), 'yellow'), colored(str(no2), 'yellow'), colored(msg1, col1), colored(text1, 'white'))
    else:
        print(colored(str(no1), 'yellow'), colored(str(no2), 'yellow'), colored(str(no3), 'yellow'), colored(msg1, col1), colored(text1, 'white'))

def get_ACP(name):
    if isinstance(acp, list):
        for each in acp:
            if each["name"]==name:
                if "id" in each.keys():
                    return {"type": each["type"], "name": each["name"], "id": each["id"]}  
                else:
                    return {"type": each["type"], "name": each["name"]}     
    else:
        if acp["name"]==name:
            if "id" in acp.keys():
                return {"type": acp["type"], "name": acp["name"], "id": acp["id"]}  
            else:
                return {"type": acp["type"], "name": acp["name"]}           

def get_NAT(name):
    if isinstance(nat, list):
        for each in nat:
            if each["name"]==name:
                if "id" in each.keys():
                    return {"type": each["type"], "name": each["name"], "id": each["id"]}  
                else:
                    return {"type": each["type"], "name": each["name"]} 
    else:
        if nat["name"]==name:
            if "id" in nat.keys():
                return {"type": nat["type"], "name": nat["name"], "id": nat["id"]}  
            else:
                return {"type": nat["type"], "name": nat["name"]}         
def get_Int(name):
    for each in Interfaces:
        if each["name"]==name:
            if "id" in each.keys():
                return {"type": each["type"], "name": each["name"], "id": each["id"]}  
            else:
                return {"type": each["type"], "name": each["name"]} 

def get_SecurityZone(name):
    for each in SecurityZones:
        if each["name"]==name:
            if "id" in each.keys():
                return {"type": each["type"], "name": each["name"], "id": each["id"]}  
            else:
                return {"type": each["type"], "name": each["name"]}      
    return None
    
def get_Object(name):
    for each in NetHostObject:
        if each["name"]==name:
            if "id" in each.keys():
                return {"type": each["type"], "name": each["name"], "id": each["id"]}
            else:
                return {"type": each["type"], "name": each["name"]}
    for each in NetworkGroups:
        if each["name"]==name:
            if "id" in each.keys():
                return {"type": each["type"], "name": each["name"], "id": each["id"]}         
            else:
                return {"type": each["type"], "name": each["name"]}    
    return None    

def get_ProtocolPortObject(name):
    for each in ProtocolPortObject:
        if each["name"]==name:
            if "id" in each.keys():
                return {"type": each["type"], "name": each["name"], "id": each["id"]}
            else:
                return {"type": each["type"], "name": each["name"]}
    for each in PortObjectGroups:
        if each["name"]==name:
            if "id" in each.keys():
                return {"name": each["name"], "type": each["type"], "overridable": each["overridable"], "id": each["id"]}     
            else:
                return {"name": each["name"], "type": each["type"], "overridable": each["overridable"]}          
    return None 

file_excel="cloudSF_PDO.xlsx"

file_FTDv_json="FTDv.json"
file_Interfaces_json="Interfaces.json"
file_SecurityZone_json="SecurityZones.json"
file_ACP_json="ACP.json"
file_ACE_json="ACE_PDO-DC-FTD-ACP_NEW.json"
file_NAT_json="NAT.json"
file_NAT_Rule_json="NAT_Entries.json"
file_NetHostObject_json= "NetHostObject.json"
file_ProtocolPortObject_json= "ProtocolPortObject.json"
file_NetworkGroups_json= "NetworkGroups.json"
file_PortObjectGroups_json= "PortObjectGroups.json"
file_PolicyAssignment="PolicyAssignment.json"

print_colored("START: ", "blue", "Script will poll the configuration from excel and save it to JSON files ",  1)

if os.path.exists(file_excel):
    wb = load_workbook(filename=file_excel)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_excel, 2)
    exit()


ws_Device = wb['Devices']
ws_Interfaces = wb['Interfaces']
ws_SecurityZones = wb['SecurityZones']   
ws_acp = wb['ACP']
ws_ace = wb['ACE']
ws_nat = wb['NAT']
ws_nat_rules = wb['NAT_Rules']
ws_Objects = wb['Objects']
ws_NetworkGroups = wb['NetworkGroup']
ws_ProtocolPortObject = wb['ProtocolPortObject']
ws_PortObjectGroups = wb['PortObjectGroups']


print_colored('PASS', 'green', "Successfully loaded the excel file: "+file_excel, 2)

##################################################################################
# we need to start with order:
# 1 ACP, NAT
# 2 Device, Interface, 
# 3 PolicyAssignment
# 4 Security Zones
# 5 Net/Host objects and port objects
# 6 Net/Port Goups
# 7 ACE/ NAT_Rules
 
#ws_acp
#A-name	B-type	C-id


row=2
lista=[]
while row <= ws_acp.max_row:
    if ws_acp["A"+str(row)].value == None:
        break
    if ws_acp["D"+str(row)].value == None:    
        lista.append({'name':ws_acp["A"+str(row)].value, 'type':ws_acp["B"+str(row)].value, 'defaultAction': {"action": ws_acp["C"+str(row)].value}})
    else:
        lista.append({'name':ws_acp["A"+str(row)].value, 'type':ws_acp["B"+str(row)].value, 'defaultAction': {"action": ws_acp["C"+str(row)].value}, 'id':ws_acp["D"+str(row)].value})
        
    row=row+1  

with open(file_ACP_json, 'w') as outfile:
    json.dump(lista[0], outfile, indent = 4)

if os.path.exists(file_ACP_json):
    with open(file_ACP_json) as json_file:
        acp=json.load(json_file)

print_colored('PASS', 'green', "Successfully saved the file: "+file_ACP_json, 3) 

#ws_nat
#A-name	B-type	C-id


row=2
lista=[]
while row <= ws_nat.max_row:
    if ws_nat["A"+str(row)].value == None:
        break
    if ws_nat["C"+str(row)].value == None:    
        lista.append({'name':ws_nat["A"+str(row)].value, 'type':ws_nat["B"+str(row)].value})
    else:
        lista.append({'name':ws_nat["A"+str(row)].value, 'type':ws_nat["B"+str(row)].value, 'id':ws_nat["C"+str(row)].value})
        
    row=row+1  


with open(file_NAT_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_NAT_json):
    with open(file_NAT_json) as json_file:
        nat=json.load(json_file)

print_colored('PASS', 'green', "Successfully saved the file: "+file_NAT_json, 4) 

#ws_Device
#name	type	id
row=2
lista=[]
Policy_Assign=[]
while row <= ws_Device.max_row:
    dictionary={}
    if ws_Device["A"+str(row)].value == None:
        print()
        break    
    dictionary["name"]=ws_Device["A"+str(row)].value
    dictionary["hostName"]=ws_Device["B"+str(row)].value
    dictionary["regKey"]=ws_Device["C"+str(row)].value
    dictionary["type"]=ws_Device["D"+str(row)].value
    dictionary["accessPolicy"]=get_ACP(ws_Device["E"+str(row)].value)
    lic_obj=[]
    for lic in str(ws_Device["G"+str(row)].value).split(", "):
        if lic != "":
            lic_obj.append(lic)
    dictionary["license_caps"]=lic_obj
    if ws_Device["H"+str(row)].value != None:
        dictionary["id"]=ws_Device["H"+str(row)].value
    lista.append(dictionary)

    if ws_Device["E"+str(row)].value != None:
        PolicyType=get_ACP(ws_Device["E"+str(row)].value)
        if ws_Device["H"+str(row)].value == None:
            TargetDevice=[{"type" : ws_Device["D"+str(row)].value, "name" : ws_Device["A"+str(row)].value }]
        else:
            TargetDevice=[{"type" : ws_Device["D"+str(row)].value, "name" : ws_Device["A"+str(row)].value,"id" : ws_Device["H"+str(row)].value }]
        Policy_Assign.append({"type": "PolicyAssignment", "policy": PolicyType, "targets": TargetDevice, "name":ws_Device["E"+str(row)].value })
    if ws_Device["F"+str(row)].value != None:
        PolicyType=get_NAT(ws_Device["F"+str(row)].value)
        if ws_Device["H"+str(row)].value == None:
            TargetDevice=[{"type" : ws_Device["D"+str(row)].value, "name" : ws_Device["A"+str(row)].value }]
        else:
            TargetDevice=[{"type" : ws_Device["D"+str(row)].value, "name" : ws_Device["A"+str(row)].value,"id" : ws_Device["H"+str(row)].value } ]      
        
        Policy_Assign.append({"type": "PolicyAssignment", "policy": PolicyType, "targets": TargetDevice, "name":ws_Device["F"+str(row)].value })
    row=row+1  


with open(file_PolicyAssignment, 'w') as outfile:
    json.dump(Policy_Assign, outfile, indent = 4)

print_colored('PASS', 'green', "Successfully saved the file: "+file_PolicyAssignment, 5) 

with open(file_FTDv_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_FTDv_json):
    with open(file_FTDv_json) as json_file:
        ftd=json.load(json_file)    


print_colored('PASS', 'green', "Successfully saved the file: "+file_FTDv_json, 6) 

#ws_Interface
#name	type	interfaceMode	enable	ipv4	dhcpRouteMetric	enableDefaultRouteDHCP	id

row=2
lista=[]
while row <= ws_Interfaces.max_row:
    int_obj={}
    if ws_Interfaces["A"+str(row)].value == None:
        break
   
    int_obj["name"]=ws_Interfaces["A"+str(row)].value
    int_obj["ifname"]=ws_Interfaces["B"+str(row)].value
    int_obj["type"]=ws_Interfaces["C"+str(row)].value
    if ws_Interfaces["D"+str(row)].value != None:
        int_obj["mode"]=ws_Interfaces["D"+str(row)].value
    if ws_Interfaces["E"+str(row)].value != None:    
        int_obj["enabled"]=bool(ws_Interfaces["E"+str(row)].value)
    if ws_Interfaces["F"+str(row)].value == "dhcp":
        int_obj["ipv4"]={"dhcp": { "dhcpRouteMetric": ws_Interfaces["G"+str(row)].value, "enableDefaultRouteDHCP": bool(ws_Interfaces["H"+str(row)].value)}}  
    if ws_Interfaces["I"+str(row)].value != None:
        int_obj["id"]=ws_Interfaces["I"+str(row)].value

    lista.append(int_obj)

    row=row+1  

with open(file_Interfaces_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)


if os.path.exists(file_Interfaces_json):
    with open(file_Interfaces_json) as json_file:
        Interfaces=json.load(json_file)

print_colored('PASS', 'green', "Successfully saved the file: "+file_Interfaces_json, 7)

#ws_SecurityZones
#name	type	id


row=2
lista=[]
int_obj=[]
sec_zone_obj={}
StartToBuild=True
while row <= ws_SecurityZones.max_row+1:
    if StartToBuild:
        if ws_SecurityZones["A"+str(row)].value != None:
            if ws_SecurityZones["E"+str(row)].value != None:
                desc=ws_SecurityZones["E"+str(row)].value
            else:
                desc=" "    
            if ws_SecurityZones["E"+str(row)].value == None:        
                each={'name': ws_SecurityZones["A"+str(row)].value, 'type': ws_SecurityZones["B"+str(row)].value, 'description' : desc, 'interfaceMode': 'ROUTED'}
            else:
                each={'name': ws_SecurityZones["A"+str(row)].value, 'type': ws_SecurityZones["B"+str(row)].value, 'description' : desc, 'interfaceMode': 'ROUTED', 'id': ws_SecurityZones["E"+str(row)].value}                            
            
            if ws_SecurityZones["C"+str(row)].value != None:
                int_obj.append(get_Int(ws_SecurityZones["C"+str(row)].value))
            StartToBuild=False
        else:
            break
    else:
        if ws_SecurityZones["A"+str(row)].value != None:
            if len(int_obj) > 0:
                each["interfaces"]=int_obj
            lista.append(each)
            each={}
            int_obj=[]

            if ws_SecurityZones["E"+str(row)].value != None:
                desc=ws_SecurityZones["E"+str(row)].value
            else:
                desc=" "    
            if ws_SecurityZones["E"+str(row)].value == None:        
                each={'name': ws_SecurityZones["A"+str(row)].value, 'type': ws_SecurityZones["B"+str(row)].value, 'description' : desc, 'interfaceMode': 'ROUTED'}
            else:
                each={'name': ws_SecurityZones["A"+str(row)].value, 'type': ws_SecurityZones["B"+str(row)].value, 'description' : desc, 'interfaceMode': 'ROUTED', 'id': ws_SecurityZones["E"+str(row)].value}                            
            
            if ws_SecurityZones["C"+str(row)].value != None:
                int_obj.append(get_Int(ws_SecurityZones["C"+str(row)].value))

        else:
            if ws_SecurityZones["C"+str(row)].value != None:
                int_obj.append(get_Int(ws_SecurityZones["C"+str(row)].value))
            else:
                if len(int_obj) > 0:
                    each["interfaces"]=int_obj
                lista.append(each)
                break
 
    row=row+1

with open(file_SecurityZone_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)
 
if os.path.exists(file_SecurityZone_json):
    with open(file_SecurityZone_json) as json_file:
        SecurityZones=json.load(json_file)

print_colored('PASS', 'green', "Successfully saved the file: "+file_SecurityZone_json, 8)
 

#Object
#A-name	B-type	C-value	D-description	E-overridable	F-id


row=2
lista=[]
while row <= ws_Objects.max_row:
    if ws_Objects["A"+str(row)].value == None:
        break
    if ws_Objects["D"+str(row)].value != None:
        desc=ws_Objects["D"+str(row)].value
    else:
        desc=" "    

    if ws_Objects["F"+str(row)].value == None:
        lista.append({'name':ws_Objects["A"+str(row)].value, 'type':ws_Objects["B"+str(row)].value, 'value':ws_Objects["C"+str(row)].value, 'description': desc, 'overridable':bool(ws_Objects["E"+str(row)].value)})
    else:
        lista.append({'name':ws_Objects["A"+str(row)].value, 'type':ws_Objects["B"+str(row)].value, 'value':ws_Objects["C"+str(row)].value, 'description': desc, 'overridable':bool(ws_Objects["E"+str(row)].value), 'id':ws_Objects["F"+str(row)].value})

     
    row=row+1

with open(file_NetHostObject_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_NetHostObject_json):
    with open(file_NetHostObject_json) as json_file:
        NetHostObject=json.load(json_file)
        
print_colored('PASS', 'green', "Successfully saved "+str(len(NetHostObject))+" objects to the file: "+file_NetHostObject_json, 9) 


#ws_ProtocolPortObject
#A-name	B-type	C-protocol	D-port	E-description	F-overridable	G-id
if os.path.exists(file_ProtocolPortObject_json):
    with open(file_ProtocolPortObject_json) as json_file:
        ProtocolPortObject=json.load(json_file)

row=2
lista=[]
while row <= ws_ProtocolPortObject.max_row:
    if ws_ProtocolPortObject["A"+str(row)].value == None:
        break
    if ws_ProtocolPortObject["E"+str(row)].value != None:
        desc=ws_ProtocolPortObject["E"+str(row)].value
    else:
        desc=" " 

    if ws_ProtocolPortObject["G"+str(row)].value == None:
        lista.append({'name': ws_ProtocolPortObject["A"+str(row)].value, 'type': ws_ProtocolPortObject["B"+str(row)].value, 'protocol': ws_ProtocolPortObject["C"+str(row)].value, 'port': ws_ProtocolPortObject["D"+str(row)].value, 'description': desc, 'overridable': bool(ws_ProtocolPortObject["F"+str(row)].value)})
    else:    
        lista.append({'name': ws_ProtocolPortObject["A"+str(row)].value, 'type': ws_ProtocolPortObject["B"+str(row)].value, 'protocol': ws_ProtocolPortObject["C"+str(row)].value, 'port': ws_ProtocolPortObject["D"+str(row)].value, 'description': desc, 'overridable': bool(ws_ProtocolPortObject["F"+str(row)].value), 'id': ws_ProtocolPortObject["G"+str(row)].value})

    row=row+1

with open(file_ProtocolPortObject_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_ProtocolPortObject_json):
    with open(file_ProtocolPortObject_json) as json_file:
        ProtocolPortObject=json.load(json_file)
 
print_colored('PASS', 'green', "Successfully saved "+str(len(ProtocolPortObject))+" objects to the file: "+file_ProtocolPortObject_json, 10) 


#ws_NetworkGroups
#A-name	B-type	C-literals	D-literals_type	E-literals_value	F-objects	G-object_name	H-description	I-overridable	J-id

if os.path.exists(file_NetworkGroups_json):
    with open(file_NetworkGroups_json) as json_file:
        NetworkGroups=json.load(json_file) 

row=2
lista=[]
StartToBuild=True
each={}
literals=[]
objects=[]
while row <= ws_NetworkGroups.max_row+1:
    if ws_NetworkGroups["A"+str(row)].value == None and ws_NetworkGroups["C"+str(row)].value == None and ws_NetworkGroups["F"+str(row)].value == None :
        if not StartToBuild:
            if len(literals) !=0:
                each["literals"]=literals
            if len(objects) !=0:
                each["objects"]=objects
            lista.append(each)    
        break
    if ws_NetworkGroups["A"+str(row)].value != None:
        if not StartToBuild:
            if len(literals) !=0:
                each["literals"]=literals
            if len(objects) !=0:
                each["objects"]=objects
            literals=[]
            objects=[]
            lista.append(each)  
            #print(each)
            each={}
            StartToBuild=True
        if ws_NetworkGroups["J"+str(row)].value == None:
            each={'name':ws_NetworkGroups["A"+str(row)].value, 'type':ws_NetworkGroups["B"+str(row)].value, 'description':ws_NetworkGroups["H"+str(row)].value, 'overridable':bool(ws_NetworkGroups["I"+str(row)].value)}
        else:
            each={'name':ws_NetworkGroups["A"+str(row)].value, 'type':ws_NetworkGroups["B"+str(row)].value, 'description':ws_NetworkGroups["H"+str(row)].value, 'overridable':bool(ws_NetworkGroups["I"+str(row)].value), 'id':ws_NetworkGroups["J"+str(row)].value}
        if ws_NetworkGroups["C"+str(row)].value != None:
            literals.append({"type":ws_NetworkGroups["D"+str(row)].value, "value": ws_NetworkGroups["E"+str(row)].value})
        if ws_NetworkGroups["F"+str(row)].value != None:
            objects.append(get_Object(ws_NetworkGroups["G"+str(row)].value)) 
        StartToBuild=False    
    else:
        if ws_NetworkGroups["C"+str(row)].value != None:
            literals.append({"type":ws_NetworkGroups["D"+str(row)].value, "value": ws_NetworkGroups["E"+str(row)].value})
        if ws_NetworkGroups["F"+str(row)].value != None:
            objects.append(get_Object(ws_NetworkGroups["G"+str(row)].value))  
    row=row+1
    
with open(file_NetworkGroups_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_NetworkGroups_json):
    with open(file_NetworkGroups_json) as json_file:
        NetworkGroups=json.load(json_file) 


print_colored('PASS', 'green', "Successfully saved "+str(len(NetworkGroups))+" objects to the file: "+file_NetworkGroups_json, 11) 
                


#ws_PortObjectGroups
#A-name	B-type	C-object_name	D-description	E-overridable	F-id

row=2
lista=[]
StartToBuild=True
each={}
objects=[]

while row <= ws_PortObjectGroups.max_row+1:
    if ws_PortObjectGroups["C"+str(row)].value == None:
        if not StartToBuild:
            if len(objects) !=0:
                each["objects"]=objects
                objects=[]
                lista.append(each)  
                each={}
                StartToBuild=True        
        break
 
    if ws_PortObjectGroups["A"+str(row)].value != None:
        if not StartToBuild:
            if len(literals) !=0:
                each["literals"]=literals
            if len(objects) !=0:
                each["objects"]=objects
            literals=[]
            objects=[]
            lista.append(each)  
            #print(each)
            each={}
            StartToBuild=True
        if ws_PortObjectGroups["F"+str(row)].value == None:
            each={'name':ws_PortObjectGroups["A"+str(row)].value, 'type':ws_PortObjectGroups["B"+str(row)].value, 'description':ws_PortObjectGroups["D"+str(row)].value, 'overridable':bool(ws_PortObjectGroups["E"+str(row)].value)}
        else:
            each={'name':ws_PortObjectGroups["A"+str(row)].value, 'type':ws_PortObjectGroups["B"+str(row)].value, 'description':ws_PortObjectGroups["D"+str(row)].value, 'overridable':bool(ws_PortObjectGroups["E"+str(row)].value), 'id':ws_PortObjectGroups["F"+str(row)].value}
        if ws_PortObjectGroups["C"+str(row)].value != None:
            objects.append(get_ProtocolPortObject(ws_PortObjectGroups["C"+str(row)].value)) 
        StartToBuild=False    
    else:
        if ws_PortObjectGroups["C"+str(row)].value != None:
            objects.append(get_ProtocolPortObject(ws_PortObjectGroups["C"+str(row)].value))  
    row=row+1

with open(file_PortObjectGroups_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

if os.path.exists(file_PortObjectGroups_json):
    with open(file_PortObjectGroups_json) as json_file:
        PortObjectGroups=json.load(json_file)   

print_colored('PASS', 'green', "Successfully saved "+str(len(PortObjectGroups))+" objects to the file: "+file_PortObjectGroups_json, 12)
  

#ws_ace
#A-ruleIndex	B-section	C-category D-name	E-sourceZones	F-destinationZones	G-sourceNetworks	H-destinationNetworks	I-destinationPorts	J-Action	K-ipsPolicy	L-variableSet	M-logBegin	N-logEnd	O-enabled	P-id


row=2
lista=[]
StartToBuild=True
each={}
sourceZones=[]
destinationZones=[]	
sourceNetworks=[]	
destinationNetworks=[]	
destinationPorts=[]

while row <= ws_ace.max_row+1:
    if ws_ace["B"+str(row)].value == None and ws_ace["E"+str(row)].value == None and ws_ace["F"+str(row)].value == None and ws_ace["E"+str(row)].value == None and ws_ace["F"+str(row)].value == None and ws_ace["G"+str(row)].value == None :
        #end of the parsing
        if not StartToBuild:
            if len(sourceZones) > 0:
                each["sourceZones"]={"objects": sourceZones}
            if len(destinationZones) > 0:
                each["destinationZones"]={"objects": destinationZones}
            if len(sourceNetworks) > 0:
                each["sourceNetworks"]={"objects": sourceNetworks}
            if len(destinationNetworks) > 0:
                each["destinationNetworks"]={"objects": destinationNetworks}
            if len(destinationPorts) > 0:
                each["destinationPorts"]={"objects": destinationPorts}       
            lista.append(each)    
        break
    if ws_ace["B"+str(row)].value != None:
        #next line in the excel 
        if not StartToBuild:
            if len(sourceZones) > 0:
                each["sourceZones"]={"objects": sourceZones}
            if len(destinationZones) > 0:
                each["destinationZones"]={"objects": destinationZones}
            if len(sourceNetworks) > 0:
                each["sourceNetworks"]={"objects": sourceNetworks}
            if len(destinationNetworks) > 0:
                each["destinationNetworks"]={"objects": destinationNetworks}
            if len(destinationPorts) > 0:
                each["destinationPorts"]={"objects": destinationPorts} 
            sourceZones=[]
            destinationZones=[]	
            sourceNetworks=[]	
            destinationNetworks=[]	
            destinationPorts=[]
            lista.append(each)  
            #print(each)
            each={}
            StartToBuild=True
        each={'name':ws_ace["B"+str(row)].value, 'action':ws_ace["H"+str(row)].value, 'logBegin':ws_ace["K"+str(row)].value, 'logEnd':ws_ace["L"+str(row)].value, 'enabled':bool(ws_ace["M"+str(row)].value)}
        if ws_ace["N"+str(row)].value != None:
            each["id"]=ws_ace["N"+str(row)].value
        if bool(ws_ace["K"+str(row)].value) or bool(ws_ace["L"+str(row)].value):
            each['sendEventsToFMC']=bool("true")
        
        
        if ws_ace["C"+str(row)].value != None:
            SecurityZone_obj=get_SecurityZone(ws_ace["C"+str(row)].value)
            if SecurityZone_obj != None:
                sourceZones.append(SecurityZone_obj)
        if ws_ace["D"+str(row)].value != None:
            destinationZones_obj=get_SecurityZone(ws_ace["D"+str(row)].value)
            if destinationZones_obj != None:
                destinationZones.append(destinationZones_obj) 
        if ws_ace["E"+str(row)].value != None:
            sourceNetworks_obj=get_Object(ws_ace["E"+str(row)].value)
            if sourceNetworks_obj != None:
                sourceNetworks.append(sourceNetworks_obj) 
        if ws_ace["F"+str(row)].value != None:
            destinationNetworks_obj=get_Object(ws_ace["F"+str(row)].value)
            if destinationNetworks_obj != None:
                destinationNetworks.append(destinationNetworks_obj) 
        if ws_ace["G"+str(row)].value != None:
            destinationPorts_obj=get_ProtocolPortObject(ws_ace["G"+str(row)].value)
            if destinationPorts_obj != None:
                destinationPorts.append(destinationPorts_obj)    

        StartToBuild=False    
    else:
        if ws_ace["C"+str(row)].value != None:
            SecurityZone_obj=get_SecurityZone(ws_ace["C"+str(row)].value)
            if SecurityZone_obj != None:
                sourceZones.append(SecurityZone_obj)
        if ws_ace["D"+str(row)].value != None:
            destinationZones_obj=get_SecurityZone(ws_ace["D"+str(row)].value)
            if destinationZones_obj != None:
                destinationZones.append(destinationZones_obj) 
        if ws_ace["E"+str(row)].value != None:
            sourceNetworks_obj=get_Object(ws_ace["E"+str(row)].value)
            if sourceNetworks_obj != None:
                sourceNetworks.append(sourceNetworks_obj) 
        if ws_ace["F"+str(row)].value != None:
            destinationNetworks_obj=get_Object(ws_ace["F"+str(row)].value)
            if destinationNetworks_obj != None:
                destinationNetworks.append(destinationNetworks_obj) 
        if ws_ace["G"+str(row)].value != None:
            destinationPorts_obj=get_ProtocolPortObject(ws_ace["G"+str(row)].value)
            if destinationPorts_obj != None:
                destinationPorts.append(destinationPorts_obj)  
    row=row+1
    
with open(file_ACE_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)

print_colored('PASS', 'green', "Successfully saved the file: "+file_ACE_json, 13) 


#NAT Rules
#enabled	type	natType	sourceInterface	destinationInterface	originalSource	interfaceInTranslatedSource	interfaceInOriginalDestination	translatedDestination	
# originalDestinationPort	translatedDestinationPort	dns	noProxyArp	routeLookup	id

row=2
lista=[]
while row <= ws_nat_rules.max_row:
    if ws_nat_rules["A"+str(row)].value == None:
        break

    if ws_nat_rules["O"+str(row)].value == None:
        lista.append({'enabled': bool(ws_nat_rules["A"+str(row)].value), 'type':ws_nat_rules["B"+str(row)].value, 'natType':ws_nat_rules["C"+str(row)].value, 'sourceInterface': get_SecurityZone(ws_nat_rules["D"+str(row)].value), 'destinationInterface':get_SecurityZone(ws_nat_rules["E"+str(row)].value), 'originalSource': get_Object(ws_nat_rules["F"+str(row)].value), 'interfaceInTranslatedSource': bool(ws_nat_rules["G"+str(row)].value), 'interfaceInOriginalDestination': bool(ws_nat_rules["H"+str(row)].value), 'translatedDestination': get_Object(ws_nat_rules["I"+str(row)].value), 'originalDestinationPort': get_ProtocolPortObject(ws_nat_rules["J"+str(row)].value), 'translatedDestinationPort': get_ProtocolPortObject(ws_nat_rules["K"+str(row)].value), 'dns': bool(ws_nat_rules["L"+str(row)].value), 'noProxyArp': bool(ws_nat_rules["M"+str(row)].value), 'routeLookup': bool(ws_nat_rules["N"+str(row)].value) })
    else:
        lista.append({'enabled': bool(ws_nat_rules["A"+str(row)].value), 'type':ws_nat_rules["B"+str(row)].value, 'natType':ws_nat_rules["C"+str(row)].value, 'sourceInterface': get_SecurityZone(ws_nat_rules["D"+str(row)].value), 'destinationInterface':get_SecurityZone(ws_nat_rules["E"+str(row)].value), 'originalSource': get_Object(ws_nat_rules["F"+str(row)].value), 'interfaceInTranslatedSource': bool(ws_nat_rules["G"+str(row)].value), 'interfaceInOriginalDestination': bool(ws_nat_rules["H"+str(row)].value), 'translatedDestination': get_Object(ws_nat_rules["I"+str(row)].value), 'originalDestinationPort': get_ProtocolPortObject(ws_nat_rules["J"+str(row)].value), 'translatedDestinationPort': get_ProtocolPortObject(ws_nat_rules["K"+str(row)].value), 'dns': bool(ws_nat_rules["L"+str(row)].value), 'noProxyArp': bool(ws_nat_rules["M"+str(row)].value), 'routeLookup': bool(ws_nat_rules["N"+str(row)].value), 'id': ws_nat_rules["O"+str(row)].value })

    row=row+1

with open(file_NAT_Rule_json, 'w') as outfile:
    json.dump(lista, outfile, indent = 4)


        
print_colored('PASS', 'green', "Successfully saved the file: "+file_NAT_Rule_json, 14) 

print_colored("FINISHED! ", "blue", "Script completed successfully!", 15)