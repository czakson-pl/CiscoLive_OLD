import os.path
from termcolor import colored
import json
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import quote_sheetname
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


file_blank_excel="cloudSF_blank.xlsx"
file_excel="cloudSF_PDO.xlsx"

file_FTDv_json="FTDv.json"
file_Interfaces_json="Interfaces.json"
file_SecurityZone_json="SecurityZones.json"
file_ACP_json="ACP.json"
file_ACE_json="ACE_PDO-DC-FTD-ACP_NEW.json"
file_NAT_json="NAT.json"
file_NAT_Rule_json="NAT_Entries.json"
file_NetHostObject_json= "NetHostObject.json"
file_ProtocolPortObject_json= "ProtocolPortObject.json"
file_NetworkGroups_json= "NetworkGroups.json"
file_PortObjectGroups_json= "PortObjectGroups.json"
file_PolicyAssignment="PolicyAssignment.json"


def print_colored(msg1, col1, text1, no1, no2=None, no3=None):
    if no3==None:
        if no2==None:
            print(colored(str(no1), 'yellow'), colored(msg1, col1), colored(text1, 'white'))
        else:
            print(colored(str(no1), 'yellow'), colored(str(no2), 'yellow'), colored(msg1, col1), colored(text1, 'white'))
    else:
        print(colored(str(no1), 'yellow'), colored(str(no2), 'yellow'), colored(str(no3), 'yellow'), colored(msg1, col1), colored(text1, 'white'))

print_colored("START: ", "blue", "Script will parse JSON files to excel file ", 1)



thin_top_border = Border( top=Side(border_style=BORDER_THIN, color='00000000') )

if os.path.exists(file_blank_excel):
    wb = load_workbook(filename=file_blank_excel)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_blank_excel, 2)
    exit()
    
if os.path.exists(file_FTDv_json):
    with open(file_FTDv_json) as json_file:
        ftd=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_FTDv_json, 2)
    exit()

if os.path.exists(file_Interfaces_json):
    with open(file_Interfaces_json) as json_file:
        Interfaces=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_Interfaces_json, 2)
    exit()

if os.path.exists(file_SecurityZone_json):
    with open(file_SecurityZone_json) as json_file:
        SecurityZones=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_SecurityZone_json, 2)
    exit()
 
if os.path.exists(file_ACP_json):
    with open(file_ACP_json) as json_file:
        acp=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_ACP_json, 2)
    exit()

if os.path.exists(file_ACE_json):
    with open(file_ACE_json) as json_file:
        ace=json.load(json_file)
        print_colored('PASS', 'green', "Loaded ACE Objects "+str(len(ace)),  2, 1)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_ACE_json, 2)
    exit()

if os.path.exists(file_NAT_json):
    with open(file_NAT_json) as json_file:
        nat=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_NAT_json, 2)
    exit()

if os.path.exists(file_NAT_Rule_json):
    with open(file_NAT_Rule_json) as json_file:
        nat_rules=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_NAT_Rule_json, 2)
    exit()

if os.path.exists(file_NetHostObject_json):
    with open(file_NetHostObject_json) as json_file:
        NetHostObject=json.load(json_file)
        print_colored('PASS', 'green', "Loaded Network/Host Objects "+str(len(NetHostObject)),  2, 2)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_NetHostObject_json, 2)
    exit()

if os.path.exists(file_ProtocolPortObject_json):
    with open(file_ProtocolPortObject_json) as json_file:
        ProtocolPortObject=json.load(json_file)
        print_colored('PASS', 'green', "Loaded Proto/Port Objects "+str(len(ProtocolPortObject)),  2, 3)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_ProtocolPortObject_json, 2)
    exit()

if os.path.exists(file_NetworkGroups_json):
    with open(file_NetworkGroups_json) as json_file:
        NetworkGroups=json.load(json_file)
        print_colored('PASS', 'green', "Loaded Network Groups Objects "+str(len(NetworkGroups)),  2, 4)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_NetworkGroups_json, 2)
    exit()

if os.path.exists(file_PortObjectGroups_json):
    with open(file_PortObjectGroups_json) as json_file:
        PortObjectGroups=json.load(json_file)
        print_colored('PASS', 'green', "Loaded Proto/Port Group Objects "+str(len(PortObjectGroups)),  2, 5)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_ProtocolPortObject_json, 2)
    exit()

if os.path.exists(file_PolicyAssignment):
    with open(file_PolicyAssignment) as json_file:
        Policy_Assign=json.load(json_file)
else:
    print_colored("ERROR: ", "red", "Cannot find the file:" + file_PolicyAssignment, 2)
    exit()

print_colored('PASS', 'green', "Successfully loaded all JSON files",  2)

ws_Device = wb['Devices']
ws_Interfaces = wb['Interfaces']
ws_SecurityZones = wb['SecurityZones']   
ws_acp = wb['ACP']
ws_ace = wb['ACE']
ws_nat = wb['NAT']
ws_nat_rules = wb['NAT_Rules']
ws_Objects = wb['Objects']
ws_NetworkGroups = wb['NetworkGroup']
ws_ProtocolPortObject = wb['ProtocolPortObject']
ws_PortObjectGroups = wb['PortObjectGroups']


dv_False_True = DataValidation(type="list", formula1="{0}!$A$2:$A$3".format(quote_sheetname("Help")), allow_blank=False)  
#Device sheet
row=2
if isinstance(ftd, list):
    for each in ftd:
        #if not isinstance(each, str):
        ws_Device["A"+str(row)]=each["name"]
        ws_Device["B"+str(row)]=each["hostName"]
        if "regKey" in each.keys():
            ws_Device["C"+str(row)]=each["regKey"]
        ws_Device["D"+str(row)]=each["type"]
        if "name" in each["accessPolicy"].keys():
            ws_Device["E"+str(row)]=each["accessPolicy"]["name"]
        if "license_caps" in each.keys():
            lic_str=""
            for lic in each["license_caps"]:
                lic_str=str(lic)+", "+lic_str
            ws_Device["G"+str(row)]=lic_str
        if "id" in each.keys():
            ws_Device["H"+str(row)]=each["id"]

        if isinstance(Policy_Assign, list):
            for each in Policy_Assign:
                if each["policy"]["type"]=="FTDNatPolicy":
                    ws_Device["F"+str(row)]=each["policy"]["name"]
        else:
            if Policy_Assign["policy"]["type"]=="FTDNatPolicy":
                ws_Device["F"+str(row)]=Policy_Assign["policy"]["name"] 

        row=row+1       
else:
    if len(ftd) > 0:
        ws_Device["A"+str(row)]=ftd["name"]
        ws_Device["B"+str(row)]=ftd["hostName"]
        ws_Device["C"+str(row)]=ftd["regKey"]
        ws_Device["D"+str(row)]=ftd["type"]
        if "name" in ftd["accessPolicy"].keys():
            ws_Device["E"+str(row)]=ftd["accessPolicy"]["name"]
        lic_str=""
        for lic in ftd["license_caps"]:
            lic_str=str(lic)+", "+lic_str
        ws_Device["G"+str(row)]=lic_str
        if "id" in ftd.keys():
            ws_Device["H"+str(row)]=ftd["id"]    
    #Device - NAT assignement check:
        
        if isinstance(Policy_Assign, list):
            for each in Policy_Assign:
                if each["policy"]["type"]=="FTDNatPolicy":
                    ws_Device["F"+str(row)]=each["policy"]["name"]
        else:
            if Policy_Assign["policy"]["type"]=="FTDNatPolicy":
                ws_Device["F"+str(row)]=Policy_Assign["policy"]["name"] 

print_colored('PASS', 'green', "Successfully parsed Device sheet",  3)

#Interfaces sheet
row=2
if isinstance(Interfaces, list):
    for each in Interfaces:
        #if not isinstance(each, str):
        ws_Interfaces["A"+str(row)]=each["name"]
        ws_Interfaces["B"+str(row)]=each["ifname"]
        ws_Interfaces["C"+str(row)]=each["type"]
        if "mode" in each.keys():
            ws_Interfaces["D"+str(row)]=each["mode"]
        if "enabled" in each.keys():
            ws_Interfaces["E"+str(row)]=each["enabled"]
        if "ipv4" in each.keys():
            if "dhcp" in each["ipv4"].keys():
                ws_Interfaces["F"+str(row)]="dhcp"
                ws_Interfaces["G"+str(row)]=str(each["ipv4"]["dhcp"]["dhcpRouteMetric"])
                if "enableDefaultRouteDHCP" in each["ipv4"]["dhcp"].keys():
                    ws_Interfaces["H"+str(row)]=each["ipv4"]["dhcp"]["enableDefaultRouteDHCP"]
        if "id" in each.keys():
            ws_Interfaces["I"+str(row)]=each["id"]
        row=row+1       
else:
    ws_Interfaces["A"+str(row)]=Interfaces["name"]
    ws_Interfaces["B"+str(row)]=Interfaces["ifname"]
    ws_Interfaces["C"+str(row)]=Interfaces["type"]
    if "mode" in Interfaces.keys():
        ws_Interfaces["D"+str(row)]=Interfaces["mode"]
    if "enabled" in Interfaces.keys():
        ws_Interfaces["E"+str(row)]=Interfaces["enabled"]
    if "ipv4" in Interfaces.keys():
        if "dhcp" in Interfaces["ipv4"].keys():
            ws_Interfaces["F"+str(row)]="dhcp"
            ws_Interfaces["G"+str(row)]=str(Interfaces["ipv4"]["dhcp"]["dhcpRouteMetric"])
            if "enableDefaultRouteDHCP" in Interfaces["ipv4"]["dhcp"].keys():
                ws_Interfaces["H"+str(row)]=Interfaces["ipv4"]["dhcp"]["enableDefaultRouteDHCP"]
    if "id" in Interfaces.keys():
        ws_Interfaces["I"+str(row)]=Interfaces["id"]    

dv_Interfaces = DataValidation(type="list", formula1="{0}!$A$2:$A${1}".format(quote_sheetname("Interfaces"),ws_Interfaces.max_row), allow_blank=False)  

print_colored('PASS', 'green', "Successfully parsed Interfaces sheet",  4)

#SecurityZones sheet
row=2
if isinstance(SecurityZones, list):
    for each in SecurityZones:
        #if not isinstance(each, str):
        ws_SecurityZones["A"+str(row)]=each["name"]
        ws_SecurityZones["B"+str(row)]=each["type"]
        if "id" in each.keys():
            ws_SecurityZones["E"+str(row)]=each["id"]

        row_hlp1=row
        if "interfaces" in each.keys():
            for each_interface in each["interfaces"]:

                cel=ws_SecurityZones["C"+str(row_hlp1)]
                cel.value=each_interface["name"]   
                dv_Interfaces.add(cel)
                ws_SecurityZones.add_data_validation(dv_Interfaces) 

                row_hlp1=row_hlp1+1
            row=row_hlp1
        else:
            row=row+1
        ws_SecurityZones["A"+str(row)].border = thin_top_border
        ws_SecurityZones["B"+str(row)].border = thin_top_border
        ws_SecurityZones["C"+str(row)].border = thin_top_border
        ws_SecurityZones["D"+str(row)].border = thin_top_border
        ws_SecurityZones["E"+str(row)].border = thin_top_border        
else:
    if len(SecurityZones) > 0:
        ws_SecurityZones["A"+str(row)]=SecurityZones["name"]
        ws_SecurityZones["B"+str(row)]=SecurityZones["type"]
        if "id" in SecurityZones.keys():
            ws_SecurityZones["E"+str(row)]=SecurityZones["id"]
        row_hlp1=row
        if "interfaces" in each.keys():
            for each_interface in each["interfaces"]:

                cel=ws_SecurityZones["C"+str(row_hlp1)]
                cel.value=each_interface["name"]   
                dv_Interfaces.add(cel)
                ws_SecurityZones.add_data_validation(dv_Interfaces) 

                row_hlp1=row_hlp1+1
            row=row_hlp1


dv_SecurityZones = DataValidation(type="list", formula1="{0}!$A$2:$A${1}".format(quote_sheetname("SecurityZones"),ws_SecurityZones.max_row), allow_blank=False)       

dv_NetType = DataValidation(type="list", formula1="{0}!$B$2:$B$5".format(quote_sheetname("Help")), allow_blank=False)

print_colored('PASS', 'green', "Successfully parsed SecurityZone sheet",  5)

#Object sheet
row=2
if isinstance(NetHostObject, list):
    for each in NetHostObject:
        #if not isinstance(each, str):
        ws_Objects["A"+str(row)]=each["name"]
        ws_Objects["B"+str(row)]=each["type"]
        ws_Objects["C"+str(row)]=each["value"]
        ws_Objects["D"+str(row)]=each["description"]
        ws_Objects["E"+str(row)]=each["overridable"]   
        if "id" in each.keys(): 
            ws_Objects["F"+str(row)]=each["id"]
        row=row+1
else:
        ws_Objects["A"+str(row)]=NetHostObject["name"]

        ws_Objects["B"+str(row)]=each["type"]
        ws_Objects["C"+str(row)]=NetHostObject["value"]
        ws_Objects["D"+str(row)]=NetHostObject["description"]
        ws_Objects["E"+str(row)]=NetHostObject["overridable"]   
        if "id" in NetHostObject.keys(): 
            ws_Objects["F"+str(row)]=NetHostObject["id"]        

#dv_NetGrpObj = DataValidation(type="list", formula1="{0}!$A$2:$A${1}".format(quote_sheetname("Objects"),ws_Objects.max_row), allow_blank=False)    

print_colored('PASS', 'green', "Successfully parsed "+str(row-2)+" NetHostObject into sheet",  6)

#NetworkGroup sheet
row=2
if isinstance(NetworkGroups, list):
    for each in NetworkGroups:
        #if not isinstance(each, str):
        ws_NetworkGroups["A"+str(row)]=each["name"]
        ws_NetworkGroups["B"+str(row)]=each["type"]
        ws_NetworkGroups["H"+str(row)]=each["description"]
        ws_NetworkGroups["I"+str(row)]=each["overridable"]    
        if "id" in each.keys():
            ws_NetworkGroups["J"+str(row)]=each["id"]   

        ws_NetworkGroups["A"+str(row)].border = thin_top_border
        ws_NetworkGroups["B"+str(row)].border = thin_top_border
        ws_NetworkGroups["C"+str(row)].border = thin_top_border
        ws_NetworkGroups["D"+str(row)].border = thin_top_border
        ws_NetworkGroups["E"+str(row)].border = thin_top_border 
        ws_NetworkGroups["F"+str(row)].border = thin_top_border
        ws_NetworkGroups["G"+str(row)].border = thin_top_border
        ws_NetworkGroups["H"+str(row)].border = thin_top_border
        ws_NetworkGroups["I"+str(row)].border = thin_top_border   
        ws_NetworkGroups["J"+str(row)].border = thin_top_border

        row_hlp1=row
        row_hlp2=row
        if "literals" in each.keys():
            for literal in each["literals"]:
                ws_NetworkGroups["C"+str(row_hlp1)]="literals"
                ws_NetworkGroups["D"+str(row_hlp1)]=literal["type"]
                ws_NetworkGroups["E"+str(row_hlp1)]=literal["value"]
                row_hlp1=row_hlp1+1
        if "objects" in each.keys():
            for obj in each["objects"]:    
                ws_NetworkGroups["F"+str(row_hlp2)]="objects"
                ws_NetworkGroups["G"+str(row_hlp2)]=obj["name"] 
                row_hlp2=row_hlp2+1
            
        row=max(row_hlp1, row_hlp2)
else:
    ws_NetworkGroups["A"+str(row)]=NetworkGroups["name"]
    ws_NetworkGroups["B"+str(row)]=NetworkGroups["type"]
    ws_NetworkGroups["H"+str(row)]=NetworkGroups["description"]
    ws_NetworkGroups["I"+str(row)]=NetworkGroups["overridable"]    
    if "id" in NetworkGroups.keys():
        ws_NetworkGroups["J"+str(row)]=NetworkGroups["id"]   

    ws_NetworkGroups["A"+str(row)].border = thin_top_border
    ws_NetworkGroups["B"+str(row)].border = thin_top_border
    ws_NetworkGroups["C"+str(row)].border = thin_top_border
    ws_NetworkGroups["D"+str(row)].border = thin_top_border
    ws_NetworkGroups["E"+str(row)].border = thin_top_border 
    ws_NetworkGroups["F"+str(row)].border = thin_top_border
    ws_NetworkGroups["G"+str(row)].border = thin_top_border
    ws_NetworkGroups["H"+str(row)].border = thin_top_border
    ws_NetworkGroups["I"+str(row)].border = thin_top_border   
    ws_NetworkGroups["J"+str(row)].border = thin_top_border

    row_hlp1=row
    row_hlp2=row
    if "literals" in NetworkGroups.keys():
        for literal in NetworkGroups["literals"]:
            ws_NetworkGroups["C"+str(row_hlp1)]="literals"
            ws_NetworkGroups["D"+str(row_hlp1)]=literal["type"]
            ws_NetworkGroups["E"+str(row_hlp1)]=literal["value"]
            row_hlp1=row_hlp1+1
    if "objects" in NetworkGroups.keys():
        for obj in NetworkGroups["objects"]:    
            ws_NetworkGroups["F"+str(row_hlp2)]="objects"
            ws_NetworkGroups["G"+str(row_hlp2)]=obj["name"]          
            row_hlp2=row_hlp2+1

print_colored('PASS', 'green', "Successfully parsed "+str(row-2)+" NetworkGroup Objects into sheet",  7)

#ProtocolPortObject sheet

row=2
if isinstance(ProtocolPortObject, list):
    for each in ProtocolPortObject:
        #if not isinstance(each, str):
        ws_ProtocolPortObject["A"+str(row)]=each["name"]
        ws_ProtocolPortObject["B"+str(row)]=each["type"]
        ws_ProtocolPortObject["C"+str(row)]=each["protocol"]
        ws_ProtocolPortObject["D"+str(row)]=each["port"]    
        ws_ProtocolPortObject["E"+str(row)]=each["description"]
        ws_ProtocolPortObject["F"+str(row)]=each["overridable"]    
        if "id" in each.keys():
            ws_ProtocolPortObject["G"+str(row)]=each["id"]
        row=row+1   
else:
    ws_ProtocolPortObject["A"+str(row)]=ProtocolPortObject["name"]
    ws_ProtocolPortObject["B"+str(row)]=ProtocolPortObject["type"]
    ws_ProtocolPortObject["C"+str(row)]=ProtocolPortObject["protocol"]
    ws_ProtocolPortObject["D"+str(row)]=ProtocolPortObject["port"]    
    ws_ProtocolPortObject["E"+str(row)]=ProtocolPortObject["description"]
    ws_ProtocolPortObject["F"+str(row)]=ProtocolPortObject["overridable"]    
    if "id" in ProtocolPortObject.keys():
        ws_ProtocolPortObject["G"+str(row)]=ProtocolPortObject["id"]
      
    
dv_ProtocolPortObject = DataValidation(type="list", formula1="{0}!$A$2:$A${1}".format(quote_sheetname("ProtocolPortObject"),ws_ProtocolPortObject.max_row), allow_blank=False)

print_colored('PASS', 'green', "Successfully parsed "+str(row-2)+" ProtocolPortObject into sheet",  8)

#ws_PortObjectGroups sheet
#name	type	object_name	description	overridable	id

row=2
if isinstance(PortObjectGroups, list):
    for each in PortObjectGroups:
        #if not isinstance(each, str):
        ws_PortObjectGroups["A"+str(row)]=each["name"]
        ws_PortObjectGroups["B"+str(row)]=each["type"]
        ws_PortObjectGroups["D"+str(row)]=each["description"]
        ws_PortObjectGroups["E"+str(row)]=each["overridable"]  
        if "id" in each.keys():  
            ws_PortObjectGroups["F"+str(row)]=each["id"]   

        ws_PortObjectGroups["A"+str(row)].border = thin_top_border
        ws_PortObjectGroups["B"+str(row)].border = thin_top_border
        ws_PortObjectGroups["C"+str(row)].border = thin_top_border
        ws_PortObjectGroups["D"+str(row)].border = thin_top_border
        ws_PortObjectGroups["E"+str(row)].border = thin_top_border 
        ws_PortObjectGroups["F"+str(row)].border = thin_top_border


        row_hlp1=row
        if "objects" in each.keys():
            for obj in each["objects"]:    
                cel=ws_PortObjectGroups["C"+str(row_hlp1)]
                cel.value=obj["name"]    
                dv_ProtocolPortObject.add(cel)
                ws_PortObjectGroups.add_data_validation(dv_ProtocolPortObject)            
                row_hlp1=row_hlp1+1
            
        row=row_hlp1
else:
    ws_PortObjectGroups["A"+str(row)]=PortObjectGroups["name"]
    ws_PortObjectGroups["B"+str(row)]=PortObjectGroups["type"]
    ws_PortObjectGroups["D"+str(row)]=PortObjectGroups["description"]
    ws_PortObjectGroups["E"+str(row)]=PortObjectGroups["overridable"]  
    if "id" in PortObjectGroups.keys():  
        ws_PortObjectGroups["F"+str(row)]=PortObjectGroups["id"]   

    ws_PortObjectGroups["A"+str(row)].border = thin_top_border
    ws_PortObjectGroups["B"+str(row)].border = thin_top_border
    ws_PortObjectGroups["C"+str(row)].border = thin_top_border
    ws_PortObjectGroups["D"+str(row)].border = thin_top_border
    ws_PortObjectGroups["E"+str(row)].border = thin_top_border 
    ws_PortObjectGroups["F"+str(row)].border = thin_top_border


    row_hlp1=row
    if "objects" in PortObjectGroups.keys():
        for obj in PortObjectGroups["objects"]:    
            cel=ws_PortObjectGroups["C"+str(row_hlp1)]
            cel.value=obj["name"]    
            dv_ProtocolPortObject.add(cel)
            ws_PortObjectGroups.add_data_validation(dv_ProtocolPortObject)            
            row_hlp1=row_hlp1+1

print_colored('PASS', 'green', "Successfully parsed "+str(row-2)+" PortObjectGroup into sheet",  9)

#ACP sheet
row=2
if isinstance(acp, list):
    for each in acp:
        #if not isinstance(each, str):
        ws_acp["A"+str(row)]=each["name"]
        ws_acp["B"+str(row)]=each["type"]
        if "defaultAction" in each.keys():
            ws_acp["C"+str(row)]=each["defaultAction"]["action"]
        if "id" in each.keys():    
            ws_acp["D"+str(row)]=each["id"]
        row=row+1
else:
    ws_acp["A"+str(row)]=acp["name"]
    ws_acp["B"+str(row)]=acp["type"]
    if "defaultAction" in acp.keys():
        ws_acp["C"+str(row)]=acp["defaultAction"]["action"]
    if "id" in acp.keys():    
        ws_acp["D"+str(row)]=acp["id"]

print_colored('PASS', 'green', "Successfully parsed ACP sheet",  10)

#ACE sheet

row=2
if isinstance(ace, list):
    for each in ace:
        #if not isinstance(each, str):
        if "metadata" in each.keys():
            ws_ace["A"+str(row)]=each["metadata"]["ruleIndex"]
        ws_ace["B"+str(row)]=each["name"] 

        ws_ace["H"+str(row)]=each["action"]
        
        if "ipsPolicy" in each.keys():
            ws_ace["I"+str(row)]=each["ipsPolicy"]["name"]
        if "variableSet" in each.keys():   
            ws_ace["J"+str(row)]=each["variableSet"]["name"]
        ws_ace["K"+str(row)]=each["logBegin"]
        ws_ace["L"+str(row)]=each["logEnd"]
        ws_ace["M"+str(row)]=each["enabled"]
        if "id" in each.keys():
            ws_ace["N"+str(row)]=each["id"]

        ws_ace["A"+str(row)].border = thin_top_border
        ws_ace["B"+str(row)].border = thin_top_border
        ws_ace["C"+str(row)].border = thin_top_border
        ws_ace["D"+str(row)].border = thin_top_border
        ws_ace["E"+str(row)].border = thin_top_border
        ws_ace["F"+str(row)].border = thin_top_border
        ws_ace["G"+str(row)].border = thin_top_border
        ws_ace["H"+str(row)].border = thin_top_border
        ws_ace["I"+str(row)].border = thin_top_border
        ws_ace["J"+str(row)].border = thin_top_border
        ws_ace["K"+str(row)].border = thin_top_border
        ws_ace["L"+str(row)].border = thin_top_border
        ws_ace["M"+str(row)].border = thin_top_border
        ws_ace["N"+str(row)].border = thin_top_border


        row_hlp1=row
        row_hlp2=row
        row_hlp3=row
        row_hlp4=row
        row_hlp5=row
        row_hlp6=row        
        
        if "sourceZones" in each.keys():
            for SrcZone in each["sourceZones"]["objects"]:
                cel=ws_ace["C"+str(row_hlp1)]
                cel.value=SrcZone["name"]    
                dv_SecurityZones.add(cel)
                ws_ace.add_data_validation(dv_SecurityZones)  
                row_hlp1=row_hlp1+1

        if "destinationZones" in each.keys():
            for DstZone in each["destinationZones"]["objects"]:
                cel=ws_ace["D"+str(row_hlp2)]
                cel.value=DstZone["name"]    
                dv_SecurityZones.add(cel)
                ws_ace.add_data_validation(dv_SecurityZones)  
                row_hlp2=row_hlp2+1

        if "sourceNetworks" in each.keys():
            if "literals" in each["sourceNetworks"]:
                for SrcNets in each["sourceNetworks"]["literals"]:
                    ws_ace["E"+str(row_hlp3)]=SrcNets["type"]+" / "+SrcNets["value"]
                    row_hlp3=row_hlp3+1
            if "objects" in each["sourceNetworks"]:
                for SrcNets in each["sourceNetworks"]["objects"]:
                    ws_ace["E"+str(row_hlp3)]=SrcNets["name"]
                    row_hlp3=row_hlp3+1
                    
        if "destinationNetworks" in each.keys():
            if "literals" in each["destinationNetworks"]:
                for DstNets in each["destinationNetworks"]["literals"]:
                    ws_ace["F"+str(row_hlp4)]=DstNets["type"]+" / "+DstNets["value"]
                    row_hlp4=row_hlp4+1
            if "objects" in each["destinationNetworks"]:
                for DstNets in each["destinationNetworks"]["objects"]:
                    ws_ace["F"+str(row_hlp4)]=DstNets["name"]
                    row_hlp4=row_hlp4+1
                    
        if "destinationPorts" in each.keys():
            if "literals" in each["destinationPorts"].keys():
                for Port_literals in each["destinationPorts"]["literals"]:
                    ws_ace["G"+str(row_hlp5)]=str(Port_literals["protocol"])+" / "+str(Port_literals["port"])      
                    row_hlp5=row_hlp5+1
            if "objects" in each["destinationPorts"].keys():
                for Port_objs in each["destinationPorts"]["objects"]:
                    ws_ace["G"+str(row_hlp6)]=Port_objs["name"]      
                    row_hlp6=row_hlp6+1
    
        row=max(row_hlp1, row_hlp2, row_hlp3, row_hlp4, row_hlp5, row_hlp6)
else:
    if "metadata" in ace.keys():
        ws_ace["A"+str(row)]=ace["metadata"]["ruleIndex"]
    ws_ace["B"+str(row)]=ace["name"] 

    ws_ace["H"+str(row)]=ace["action"]
    
    if "ipsPolicy" in ace.keys():
        ws_ace["I"+str(row)]=ace["ipsPolicy"]["name"]
        
    ws_ace["J"+str(row)]=ace["variableSet"]["name"]
    ws_ace["K"+str(row)]=ace["logBegin"]
    ws_ace["L"+str(row)]=ace["logEnd"]
    ws_ace["M"+str(row)]=ace["enabled"]
    if "id" in ace.keys():
        ws_ace["N"+str(row)]=ace["id"]

    ws_ace["A"+str(row)].border = thin_top_border
    ws_ace["B"+str(row)].border = thin_top_border
    ws_ace["C"+str(row)].border = thin_top_border
    ws_ace["D"+str(row)].border = thin_top_border
    ws_ace["E"+str(row)].border = thin_top_border
    ws_ace["F"+str(row)].border = thin_top_border
    ws_ace["G"+str(row)].border = thin_top_border
    ws_ace["H"+str(row)].border = thin_top_border
    ws_ace["I"+str(row)].border = thin_top_border
    ws_ace["J"+str(row)].border = thin_top_border
    ws_ace["K"+str(row)].border = thin_top_border
    ws_ace["L"+str(row)].border = thin_top_border
    ws_ace["M"+str(row)].border = thin_top_border
    ws_ace["N"+str(row)].border = thin_top_border

    row_hlp1=row
    row_hlp2=row
    row_hlp3=row
    row_hlp4=row
    row_hlp5=row
    row_hlp6=row        
    
    if "sourceZones" in ace.keys():
        for SrcZone in ace["sourceZones"]["objects"]:
            cel=ws_ace["C"+str(row_hlp1)]
            cel.value=SrcZone["name"]    
            dv_SecurityZones.add(cel)
            ws_ace.add_data_validation(dv_SecurityZones)  
            row_hlp1=row_hlp1+1

    if "destinationZones" in ace.keys():
        for DstZone in ace["destinationZones"]["objects"]:
            cel=ws_ace["D"+str(row_hlp2)]
            cel.value=DstZone["name"]    
            dv_SecurityZones.add(cel)
            ws_ace.add_data_validation(dv_SecurityZones)  
            row_hlp2=row_hlp2+1

    if "sourceNetworks" in ace.keys():
        if "literals" in ace["sourceNetworks"]:
            for SrcNets in ace["sourceNetworks"]["literals"]:
                ws_ace["E"+str(row_hlp3)]=SrcNets["type"]+" / "+SrcNets["value"]
                row_hlp3=row_hlp3+1
        if "objects" in ace["sourceNetworks"]:
            for SrcNets in ace["sourceNetworks"]["objects"]:
                ws_ace["E"+str(row_hlp3)]=SrcNets["name"]
                row_hlp3=row_hlp3+1
                
    if "destinationNetworks" in ace.keys():
        if "literals" in ace["destinationNetworks"]:
            for DstNets in ace["destinationNetworks"]["literals"]:
                ws_ace["F"+str(row_hlp4)]=DstNets["type"]+" / "+DstNets["value"]
                row_hlp4=row_hlp4+1
        if "objects" in ace["destinationNetworks"]:
            for DstNets in ace["destinationNetworks"]["objects"]:
                ws_ace["F"+str(row_hlp4)]=DstNets["name"]
                row_hlp4=row_hlp4+1
                
    if "destinationPorts" in ace.keys():
        if "literals" in ace["destinationPorts"].keys():
            for Port_literals in ace["destinationPorts"]["literals"]:
                ws_ace["G"+str(row_hlp5)]=str(Port_literals["protocol"])+" / "+str(Port_literals["port"])      
                row_hlp5=row_hlp5+1
        if "objects" in ace["destinationPorts"].keys():
            for Port_objs in ace["destinationPorts"]["objects"]:
                ws_ace["G"+str(row_hlp6)]=Port_objs["name"]      
                row_hlp6=row_hlp6+1
    

print_colored('PASS', 'green', "Successfully parsed ACE sheet",  11)

#NAT sheet
row=2
if isinstance(nat, list):
    for each in nat:
        #if not isinstance(each, str):
        ws_nat["A"+str(row)]=each["name"]
        ws_nat["B"+str(row)]=each["type"]
        if "id" in each.keys():    
            ws_nat["C"+str(row)]=each["id"]
        row=row+1
else:
    if len(nat) > 0:
        ws_nat["A"+str(row)]=nat["name"]
        ws_nat["B"+str(row)]=nat["type"]
        if "id" in nat.keys():    
            ws_nat["C"+str(row)]=nat["id"]


print_colored('PASS', 'green', "Successfully parsed NAT sheet",  13)
#NAT RUles sheet

row=2
if isinstance(nat_rules, list):
    for each in nat_rules:
        ws_nat_rules["A"+str(row)]=each["enabled"] 
        ws_nat_rules["B"+str(row)]=each["type"]
        ws_nat_rules["C"+str(row)]=each["natType"]
        ws_nat_rules["D"+str(row)]=each["sourceInterface"]["name"]
        ws_nat_rules["E"+str(row)]=each["destinationInterface"]["name"]
        ws_nat_rules["F"+str(row)]=each["originalSource"]["name"]
        ws_nat_rules["G"+str(row)]=each["interfaceInTranslatedSource"]
        ws_nat_rules["H"+str(row)]=each["interfaceInOriginalDestination"]
        ws_nat_rules["I"+str(row)]=each["translatedDestination"]["name"]
        ws_nat_rules["J"+str(row)]=each["originalDestinationPort"]["name"]
        ws_nat_rules["K"+str(row)]=each["translatedDestinationPort"]["name"]
        ws_nat_rules["L"+str(row)]=each["dns"]
        ws_nat_rules["M"+str(row)]=each["noProxyArp"]
        ws_nat_rules["N"+str(row)]=each["routeLookup"]
        if "id" in each.keys():
            ws_nat_rules["O"+str(row)]=each["id"]
        row+=1

else:
    ws_nat_rules["A"+str(row)]=nat_rules["enabled"] 
    ws_nat_rules["B"+str(row)]=nat_rules["type"]
    ws_nat_rules["C"+str(row)]=nat_rules["natType"]
    ws_nat_rules["D"+str(row)]=nat_rules["sourceInterface"]["name"]
    ws_nat_rules["E"+str(row)]=nat_rules["destinationInterface"]["name"]
    ws_nat_rules["F"+str(row)]=nat_rules["originalSource"]["name"]
    ws_nat_rules["G"+str(row)]=nat_rules["interfaceInTranslatedSource"]
    ws_nat_rules["H"+str(row)]=nat_rules["interfaceInOriginalDestination"]
    ws_nat_rules["I"+str(row)]=nat_rules["translatedDestination"]["name"]
    ws_nat_rules["J"+str(row)]=nat_rules["originalDestinationPort"]["name"]
    ws_nat_rules["K"+str(row)]=nat_rules["translatedDestinationPort"]["name"]
    ws_nat_rules["L"+str(row)]=nat_rules["dns"]
    ws_nat_rules["M"+str(row)]=nat_rules["noProxyArp"]
    ws_nat_rules["N"+str(row)]=nat_rules["routeLookup"]
    if "id" in nat_rules.keys():
        ws_nat_rules["O"+str(row)]=nat_rules["id"]
    
print_colored('PASS', 'green', "Successfully parsed NAT Rules sheet",  14)


#Save file
wb.save(filename=file_excel)


print_colored("FINISHED! ", "blue", "Script completed successfully!", 15)